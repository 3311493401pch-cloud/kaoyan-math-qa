"""
考研数学知识库构建器
解析 Excel → 拆分为结构化 chunk → 向量化 → 存入 Chroma
"""
import os
import re
import hashlib
from pathlib import Path

import openpyxl
import chromadb
from chromadb.config import Settings
from sentence_transformers import SentenceTransformer

# ============================================================
# 配置
# ============================================================
EXCEL_PATH = "亚瑟爱数学考研数学规划.xlsx"
CHROMA_DIR = "./chroma_data"
EMBEDDING_MODEL = "BAAI/bge-small-zh-v1.5"
COLLECTION_NAME = "kaoyan_math"

# 规划表 Sheet → 老师名映射
PLAN_SHEETS = {
    "张宇": "张宇",
    "武忠祥A1": "武忠祥（方案A1）",
    "武忠祥A2": "武忠祥（方案A2）",
}

# 科目列映射（规划表中 B=高数, C=线代, D=概率论）
SUBJECT_COL_MAP = {
    1: "高数",
    2: "线代",
    3: "概率论",
}


def load_excel(path: str) -> openpyxl.Workbook:
    """加载 Excel，返回 workbook"""
    return openpyxl.load_workbook(path, data_only=True)


def safe_text(value) -> str:
    """把单元格值转成干净字符串"""
    if value is None:
        return ""
    text = str(value).strip()
    # 把 "\" 这种占位符去掉
    if text in ("\\", "-", "—", "/"):
        return ""
    return text


# ============================================================
# 1. 解析规划表（张宇 / 武忠祥A1 / 武忠祥A2）
# ============================================================
def parse_plan_sheet(ws, teacher_name: str) -> list[dict]:
    """
    解析规划表 Sheet。
    结构：
      Row 1: 说明 | 高数 | 线代 | 概率论 | 分数目标
      Row 2~6: 阶段名 | 高数内容 | 线代内容 | 概率论内容 | (分数合并)
      Row 7: 过渡说明（合并行）
      Row 8~9: 真题 / 模拟卷
    """
    chunks = []

    # --- 提取分数目标（E 列，第5列）---
    score_targets = safe_text(ws.cell(1, 5).value)

    # --- 提取表头说明 ---
    header_note = safe_text(ws.cell(1, 1).value)

    # --- 解析阶段行（Row 2~6）---
    for row_idx in range(2, 7):
        stage_name = safe_text(ws.cell(row_idx, 1).value)
        if not stage_name:
            continue

        # 每个科目列
        for col_idx in range(1, 4):  # B, C, D 列
            content = safe_text(ws.cell(row_idx, col_idx + 1).value)
            if not content:
                continue

            subject = SUBJECT_COL_MAP.get(col_idx, "通用")

            chunk_text = f"【{teacher_name}】{stage_name} — {subject}\n{content}"

            chunks.append({
                "id": hashlib.md5(chunk_text.encode()).hexdigest()[:16],
                "content": chunk_text,
                "metadata": {
                    "type": "规划",
                    "teacher": teacher_name,
                    "stage": stage_name.split("\n")[0],  # 只取第一行作为阶段名
                    "subject": subject,
                    "score_targets": score_targets[:200],
                    "source_sheet": ws.title,
                }
            })

    # --- 解析真题行（Row 8）---
    zhenti_content = safe_text(ws.cell(8, 2).value)  # B8（合并了 B:D）
    if zhenti_content:
        chunk_text = f"【{teacher_name}】真题阶段\n{zhenti_content}"
        chunks.append({
            "id": hashlib.md5(chunk_text.encode()).hexdigest()[:16],
            "content": chunk_text,
            "metadata": {
                "type": "规划",
                "teacher": teacher_name,
                "stage": "真题",
                "subject": "全部",
                "score_targets": score_targets[:200],
                "source_sheet": ws.title,
            }
        })

    # --- 解析模拟卷行（Row 9）---
    moni_content = safe_text(ws.cell(9, 2).value)  # B9（合并了 B:D）
    if moni_content:
        chunk_text = f"【{teacher_name}】模拟卷阶段\n{moni_content}"
        chunks.append({
            "id": hashlib.md5(chunk_text.encode()).hexdigest()[:16],
            "content": chunk_text,
            "metadata": {
                "type": "规划",
                "teacher": teacher_name,
                "stage": "模拟卷",
                "subject": "全部",
                "score_targets": score_targets[:200],
                "source_sheet": ws.title,
            }
        })

    # --- 分数目标单独一条 ---
    if score_targets:
        chunk_text = f"【{teacher_name}】分数目标与总体规划思路\n{score_targets}"
        chunks.append({
            "id": hashlib.md5(chunk_text.encode()).hexdigest()[:16],
            "content": chunk_text,
            "metadata": {
                "type": "规划",
                "teacher": teacher_name,
                "stage": "分数目标",
                "subject": "全部",
                "source_sheet": ws.title,
            }
        })

    return chunks


# ============================================================
# 2. 解析名师测评表
# ============================================================
def parse_teacher_review_sheet(ws) -> list[dict]:
    """
    结构：
      Row 1: 表头
      Row 2~13: 科目(合并) | 老师名 | 测评内容
    """
    chunks = []
    current_subject = ""

    for row_idx in range(2, ws.max_row + 1):
        # A 列：科目（合并单元格）
        subject_val = safe_text(ws.cell(row_idx, 1).value)
        if subject_val:
            current_subject = subject_val

        teacher_name = safe_text(ws.cell(row_idx, 2).value)
        review = safe_text(ws.cell(row_idx, 3).value)

        if not teacher_name or not review:
            continue

        chunk_text = f"【名师测评】{current_subject} — {teacher_name}\n{review}"

        chunks.append({
            "id": hashlib.md5(chunk_text.encode()).hexdigest()[:16],
            "content": chunk_text,
            "metadata": {
                "type": "名师测评",
                "subject": current_subject,
                "teacher": teacher_name,
                "source_sheet": ws.title,
            }
        })

    return chunks


# ============================================================
# 3. 解析书籍测评表
# ============================================================
def parse_book_review_sheet(ws) -> list[dict]:
    """
    结构：
      Row 1: 习题册测评（长文本，按书名拆）
      Row 2: 模拟卷测评（长文本，按书名拆）
    """
    chunks = []

    for row_idx in range(1, ws.max_row + 1):
        raw_text = safe_text(ws.cell(row_idx, 1).value)
        if not raw_text:
            continue

        row_type = "习题册" if row_idx == 1 else "模拟卷"

        # 用书名号「书名：」或「书名\n」来拆分
        # 常见的书名模式：XXX题、XXX卷、XXX讲义
        # 用正则按「XXX：」模式拆分
        entries = re.split(r'(?=[^\s：]+(?:题|卷|套卷|讲义|系列)[：:])', raw_text)
        # 去掉空串
        entries = [e.strip() for e in entries if e.strip()]

        if len(entries) <= 1:
            # 拆不出来就整段当作一条
            entries = [raw_text]

        for entry in entries:
            if len(entry) < 20:  # 太短跳过
                continue
            chunk_text = f"【书籍测评】{row_type}\n{entry}"

            chunks.append({
                "id": hashlib.md5(chunk_text.encode()).hexdigest()[:16],
                "content": chunk_text,
                "metadata": {
                    "type": "书籍测评",
                    "category": row_type,
                    "source_sheet": ws.title,
                }
            })

    return chunks


# ============================================================
# 4. 主流程：构建知识库
# ============================================================
def build_knowledge_base(rebuild: bool = False):
    """主入口：解析 Excel，构建 Chroma 向量库"""
    print(f"📂 加载 Excel: {EXCEL_PATH}")
    wb = load_excel(EXCEL_PATH)

    print(f"🤖 加载 Embedding 模型: {EMBEDDING_MODEL}")
    model = SentenceTransformer(EMBEDDING_MODEL)

    # 初始化 Chroma
    os.makedirs(CHROMA_DIR, exist_ok=True)
    client = chromadb.PersistentClient(path=CHROMA_DIR)

    # 如果需要重建，先删旧集合
    try:
        if rebuild:
            client.delete_collection(COLLECTION_NAME)
            print("🗑️  已删除旧集合，重建中...")
    except Exception:
        pass

    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"description": "考研数学全年规划知识库"}
    )

    # 如果已经有数据且不重建，跳过
    if collection.count() > 0 and not rebuild:
        print(f"✅ 知识库已有 {collection.count()} 条记录，跳过构建")
        return collection

    # --- 解析所有 Sheet ---
    all_chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"📄 解析 Sheet: {sheet_name}")

        if sheet_name in PLAN_SHEETS:
            chunks = parse_plan_sheet(ws, PLAN_SHEETS[sheet_name])
        elif sheet_name == "名师测评":
            chunks = parse_teacher_review_sheet(ws)
        elif sheet_name == "书籍测评":
            chunks = parse_book_review_sheet(ws)
        else:
            # 冲刺期规划 / 章节考察方法 → 空表，跳过
            continue

        all_chunks.extend(chunks)
        print(f"   → {len(chunks)} 条 chunk")

    print(f"\n📊 共 {len(all_chunks)} 条 chunk，开始向量化...")

    # 批量写入 Chroma
    batch_size = 32
    for i in range(0, len(all_chunks), batch_size):
        batch = all_chunks[i:i + batch_size]
        texts = [c["content"] for c in batch]
        ids = [c["id"] for c in batch]
        metas = [c["metadata"] for c in batch]

        embeddings = model.encode(texts, show_progress_bar=False).tolist()

        collection.add(
            ids=ids,
            embeddings=embeddings,
            documents=texts,
            metadatas=metas,
        )
        print(f"   ✅ 已写入 {min(i + batch_size, len(all_chunks))}/{len(all_chunks)}")

    print(f"\n🎉 知识库构建完成！共 {collection.count()} 条记录")
    return collection


if __name__ == "__main__":
    build_knowledge_base(rebuild=True)
